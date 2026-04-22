# Module to handle files, and run pandas
import json
from io import BytesIO

import docx
import markdown
import pandas as pd
from htmldocx import HtmlToDocx
from openpyxl.styles import PatternFill


# String to word doc
def generate_doc(contents: str) -> BytesIO:
    doc = docx.Document()
    html = markdown.markdown(contents)
    HtmlToDocx().add_html_to_document(html, doc)
    buffer = BytesIO()
    doc.save(buffer)
    buffer.seek(0)
    return buffer


HIGHLIGHT = {
    "added": "C6EFCE",  # green
    "removed": "FFC7CE",  # red
    "modified": "FFEB9C",  # yellow
}


def generate_excel(contents: str) -> BytesIO:
    rows = json.loads(contents)
    df = pd.DataFrame(rows)
    # Drop diff column before writing — it's metadata, not display data
    df_display = df.drop(columns=["diff"], errors="ignore")

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_display.to_excel(writer, index=False, sheet_name="Inventory")
        ws = writer.sheets["Inventory"]
        cols = list(df_display.columns)

        for i, row in enumerate(rows, start=2):  # row 1 is header
            change = row.get("change", "unchanged")
            diff = row.get("diff", {})
            fill_row = HIGHLIGHT.get(change)

            for j, col in enumerate(cols, start=1):
                cell = ws.cell(row=i, column=j)
                if change == "modified" and col in diff:
                    # Highlight only changed cells
                    cell.fill = PatternFill("solid", fgColor=HIGHLIGHT["modified"])
                elif fill_row and change != "modified":
                    # Highlight whole row for added/removed
                    cell.fill = PatternFill("solid", fgColor=fill_row)

    buffer.seek(0)
    return buffer
